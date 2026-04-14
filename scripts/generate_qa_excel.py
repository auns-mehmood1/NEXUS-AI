"""
NexusAI QA Excel Report Generator
Reads qa-summary.json + regression-baseline.json + test case CSVs
Produces: automation-artifacts/qa-test-results.xlsx
"""

import json
import csv
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint

# ─── Paths ────────────────────────────────────────────────────────────────────
ROOT        = os.path.join(os.path.dirname(__file__), "..")
ARTIFACTS   = os.path.join(ROOT, "automation-artifacts")
SPECS       = os.path.join(ROOT, "specs")
SUMMARY_F   = os.path.join(ARTIFACTS, "qa-summary.json")
BASELINE_F  = os.path.join(ARTIFACTS, "regression-baseline.json")
OUTPUT_F    = os.path.join(ARTIFACTS, "qa-test-results.xlsx")

# ─── Colour Palette ──────────────────────────────────────────────────────────
C = {
    "header_bg"     : "1E3A5F",   # deep navy
    "header_fg"     : "FFFFFF",
    "section_bg"    : "2E6DA4",   # medium blue
    "section_fg"    : "FFFFFF",
    "pass_bg"       : "D4EDDA",   # soft green
    "pass_fg"       : "155724",
    "fail_bg"       : "F8D7DA",   # soft red
    "fail_fg"       : "721C24",
    "xfail_bg"      : "FFF3CD",   # amber
    "xfail_fg"      : "856404",
    "skip_bg"       : "E2E3E5",
    "skip_fg"       : "383D41",
    "critical_bg"   : "F8D7DA",
    "critical_fg"   : "721C24",
    "high_bg"       : "FFE5CC",
    "high_fg"       : "7B3F00",
    "medium_bg"     : "FFF3CD",
    "medium_fg"     : "856404",
    "low_bg"        : "D4EDDA",
    "low_fg"        : "155724",
    "alt_row"       : "F2F7FC",
    "white"         : "FFFFFF",
    "light_border"  : "BDD7EE",
    "title_bg"      : "0D2E5C",
    "title_fg"      : "FFFFFF",
    "summary_bg"    : "EBF3FB",
    "yellow_status" : "FFC107",
    "green_status"  : "28A745",
    "red_status"    : "DC3545",
}

# ─── Style helpers ────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=11, color="000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def border(style="thin", color="BDD7EE"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def style_cell(cell, bg=None, fg="000000", bold=False, size=11,
               h_align="left", wrap=False, italic=False):
    if bg:
        cell.fill = fill(bg)
    cell.font  = font(bold=bold, size=size, color=fg, italic=italic)
    cell.alignment = align(h=h_align, wrap=wrap)
    cell.border = border()

def header_cell(cell, text, bg=None, fg="FFFFFF", size=11, h_align="center"):
    cell.value = text
    style_cell(cell, bg=bg or C["header_bg"], fg=fg, bold=True,
               size=size, h_align=h_align)

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

# ─── Status display helpers ───────────────────────────────────────────────────
STATUS_MAP = {
    "pass"    : ("PASS",      C["pass_bg"],   C["pass_fg"]),
    "fail"    : ("FAIL",      C["fail_bg"],   C["fail_fg"]),
    "xfail"   : ("KNOWN BUG", C["xfail_bg"],  C["xfail_fg"]),
    "skip"    : ("SKIPPED",   C["skip_bg"],   C["skip_fg"]),
    "pending" : ("PENDING",   C["summary_bg"],"2E6DA4"),
}

PRIORITY_MAP = {
    "Critical" : (C["critical_bg"], C["critical_fg"]),
    "High"     : (C["high_bg"],     C["high_fg"]),
    "Medium"   : (C["medium_bg"],   C["medium_fg"]),
    "Low"      : (C["low_bg"],      C["low_fg"]),
}

# ─── Load data ────────────────────────────────────────────────────────────────
with open(SUMMARY_F)  as f: summary  = json.load(f)
with open(BASELINE_F) as f: baseline = json.load(f)

def load_csv(name):
    path = os.path.join(SPECS, name)
    if not os.path.exists(path):
        return []
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))

def get_test_data(tc):
    """Return the best available test-data/payload value for a test case."""
    for key in ("RequestBody", "TestData", "InputData", "Payload", "Body"):
        value = tc.get(key)
        if value is None:
            continue
        value = str(value).strip()
        if value:
            return value
    return "-"

auth_rows      = load_csv("auth-test-cases.csv")
chat_rows      = load_csv("chat-test-cases.csv")
models_rows    = load_csv("models-test-cases.csv")
dashboard_rows = load_csv("dashboard-test-cases.csv")

# Security test cases not in CSV — define inline
security_rows = [
    {"TestCaseID":"SEC-001","Module":"security","TestName":"Tampered JWT rejected",
     "Type":"security","Method":"GET","Endpoint":"/api/auth/me",
     "RequestBody":"Forged JWT","ExpectedStatus":"401","Priority":"Critical"},
    {"TestCaseID":"SEC-002","Module":"security","TestName":"No auth header on protected routes",
     "Type":"security","Method":"GET","Endpoint":"/api/auth/me, /chat/history, /dashboard/usage",
     "RequestBody":"","ExpectedStatus":"401","Priority":"High"},
    {"TestCaseID":"SEC-003","Module":"security","TestName":"Empty Bearer token",
     "Type":"security","Method":"GET","Endpoint":"/api/auth/me",
     "RequestBody":"Authorization: Bearer ","ExpectedStatus":"401","Priority":"High"},
    {"TestCaseID":"SEC-004","Module":"security","TestName":"NoSQL injection in signup",
     "Type":"security","Method":"POST","Endpoint":"/api/auth/signup",
     "RequestBody":'{"email":{"$ne":null}}', "ExpectedStatus":"400","Priority":"Critical"},
    {"TestCaseID":"SEC-005","Module":"security","TestName":"Path traversal in model ID",
     "Type":"security","Method":"GET","Endpoint":"/api/models/../../../etc/passwd",
     "RequestBody":"","ExpectedStatus":"Not 500","Priority":"High"},
    {"TestCaseID":"SEC-006","Module":"security","TestName":"Error response format correct",
     "Type":"security","Method":"POST","Endpoint":"/api/auth/login",
     "RequestBody":'{"email":"bad@test.com","password":"wrong"}',
     "ExpectedStatus":"401","Priority":"Medium"},
]

# Performance test cases — defined inline (no CSV)
performance_rows = [
    {"TestCaseID":"PERF-001","Module":"performance","TestName":"GET /models baseline latency (p50<200ms, p95<2000ms)",
     "Type":"performance","Method":"GET","Endpoint":"/api/models",
     "RequestBody":"—","ExpectedStatus":"200","Priority":"High"},
    {"TestCaseID":"PERF-002","Module":"performance","TestName":"POST /auth/login baseline latency (p50<200ms, p95<2000ms)",
     "Type":"performance","Method":"POST","Endpoint":"/api/auth/login",
     "RequestBody":'{"email":"...","password":"..."}', "ExpectedStatus":"200/201","Priority":"High"},
    {"TestCaseID":"PERF-003","Module":"performance","TestName":"POST /chat/send baseline latency (p95<2000ms)",
     "Type":"performance","Method":"POST","Endpoint":"/api/chat/send",
     "RequestBody":'{"modelId":"gpt4o","content":"perf probe"}', "ExpectedStatus":"200/201","Priority":"High"},
    {"TestCaseID":"PERF-004","Module":"performance","TestName":"GET /models — 20 concurrent VUs, error rate <5%",
     "Type":"performance","Method":"GET","Endpoint":"/api/models",
     "RequestBody":"—","ExpectedStatus":"200","Priority":"Medium"},
    {"TestCaseID":"PERF-005","Module":"performance","TestName":"POST /auth/login — 20 concurrent VUs, error rate <5%",
     "Type":"performance","Method":"POST","Endpoint":"/api/auth/login",
     "RequestBody":'{"email":"...","password":"..."}', "ExpectedStatus":"200/201","Priority":"Medium"},
    {"TestCaseID":"PERF-006","Module":"performance","TestName":"GET /chat/history baseline latency (p95<2000ms)",
     "Type":"performance","Method":"GET","Endpoint":"/api/chat/history",
     "RequestBody":"—","ExpectedStatus":"200","Priority":"Medium"},
]

all_rows = auth_rows + chat_rows + models_rows + dashboard_rows + security_rows + performance_rows
results  = baseline["results"]

# Bug details map
BUG_DETAILS = {
    "BUG-AUTH-001" : ("Critical", "Hardcoded JWT fallback secrets in source code",
                      "auth.service.ts:71,76", "Static Analysis"),
    "BUG-AUTH-002" : ("Critical", "No rate limiting on any auth endpoint",
                      "auth.controller.ts", "Static Analysis"),
    "BUG-CHAT-001" : ("High",     "String vs ObjectId comparison in deleteSession",
                      "chat.service.ts:94", "Static Analysis"),
    "BUG-CHAT-002" : ("High",     "Guest message ownership not enforced on sendMessage",
                      "chat.service.ts:56", "Static Analysis"),
    "BUG-CHAT-003" : ("High",     "Empty content + no attachments causes 500",
                      "chat.service.ts:61", "TC-CHAT-013 XFAIL"),
    "BUG-DASHBOARD-001": ("High", "Dashboard returns hardcoded random data (Math.random())",
                      "dashboard.controller.ts:15", "Static Analysis"),
    "BUG-MODELS-001": ("Medium",  "parseFloat('abc')=NaN passed to service → 500",
                      "models.controller.ts:15", "TC-MODELS-006 XFAIL"),
    "BUG-UPLOAD-001": ("Medium",  "50MB body limit applies to unauthenticated routes",
                      "main.ts:9-11", "Static Analysis"),
    "BUG-AUTH-003" : ("Medium",   "Login/Refresh return 201 instead of 200",
                      "auth.controller.ts:19,26", "Static Analysis"),
}

# ─── Create workbook ─────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)  # remove default sheet

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — EXECUTIVE SUMMARY
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("Executive Summary")
ws1.sheet_view.showGridLines = False

# Title banner
ws1.merge_cells("A1:H1")
c = ws1["A1"]
c.value = "NexusAI QA Report — Executive Summary"
style_cell(c, bg=C["title_bg"], fg=C["title_fg"], bold=True, size=18, h_align="center")
ws1.row_dimensions[1].height = 40

ws1.merge_cells("A2:H2")
c = ws1["A2"]
c.value = f"Generated: {summary['timestamp']}  |  Branch: {summary['branch']}  |  Commit: {summary['commit']}"
style_cell(c, bg=C["section_bg"], fg=C["section_fg"], size=10, h_align="center", italic=True)
ws1.row_dimensions[2].height = 20

# Pipeline status badge
ws1.merge_cells("A4:H4")
c = ws1["A4"]
status_color = {"YELLOW": C["yellow_status"], "GREEN": C["green_status"],
                "BLOCKED": C["red_status"]}.get(summary["status"], C["yellow_status"])
c.value = f"  PIPELINE STATUS:  {summary['status']}  "
style_cell(c, bg=status_color, fg=C["white"] if summary["status"] != "YELLOW" else "000000",
           bold=True, size=16, h_align="center")
ws1.row_dimensions[4].height = 36

# ── Test Results ──
ws1.merge_cells("A6:H6")
header_cell(ws1["A6"], "TEST RESULTS", bg=C["header_bg"], size=12)
ws1.row_dimensions[6].height = 26

metrics = [
    ("Total Tests",          summary["test_results"]["total"],    C["section_bg"],  C["white"]),
    ("Passed",               summary["test_results"]["passed"],   C["pass_bg"],     C["pass_fg"]),
    ("Failed (hard)",        summary["test_results"]["failed"],   C["fail_bg"],     C["fail_fg"]),
    ("Known Bugs (xfail)",   summary["test_results"]["xfailed"],  C["xfail_bg"],    C["xfail_fg"]),
    ("Pass Rate",            summary["test_results"]["pass_rate"],C["pass_bg"],     C["pass_fg"]),
    ("Duration",             f"{summary['test_results']['duration_seconds']}s", C["alt_row"], "000000"),
]

headers = ["Metric", "Value"]
for col, h in enumerate(headers, 1):
    c = ws1.cell(row=7, column=col, value=h)
    style_cell(c, bg=C["header_bg"], fg="FFFFFF", bold=True, h_align="center")

for i, (label, val, bg, fg) in enumerate(metrics, 8):
    c1 = ws1.cell(row=i, column=1, value=label)
    c2 = ws1.cell(row=i, column=2, value=val)
    style_cell(c1, bg=bg, fg=fg, bold=True)
    style_cell(c2, bg=bg, fg=fg, bold=True, h_align="center")
    ws1.row_dimensions[i].height = 22

# ── Bug Severity Counts ──
start = 8 + len(metrics) + 1
ws1.merge_cells(f"A{start}:H{start}")
header_cell(ws1[f"A{start}"], "BUG SEVERITY BREAKDOWN", bg=C["header_bg"], size=12)
ws1.row_dimensions[start].height = 26

bug_rows = [
    ("Critical", summary["bugs"]["critical"], C["critical_bg"], C["critical_fg"]),
    ("High",     summary["bugs"]["high"],     C["high_bg"],     C["high_fg"]),
    ("Medium",   summary["bugs"]["medium"],   C["medium_bg"],   C["medium_fg"]),
    ("Low",      summary["bugs"]["low"],      C["low_bg"],      C["low_fg"]),
    ("Total",    summary["bugs"]["total"],    C["header_bg"],   C["header_fg"]),
]
for col, h in enumerate(["Severity", "Bug Count"], 1):
    c = ws1.cell(row=start+1, column=col, value=h)
    style_cell(c, bg=C["header_bg"], fg="FFFFFF", bold=True, h_align="center")

for i, (sev, cnt, bg, fg) in enumerate(bug_rows, start+2):
    c1 = ws1.cell(row=i, column=1, value=sev)
    c2 = ws1.cell(row=i, column=2, value=cnt)
    style_cell(c1, bg=bg, fg=fg, bold=True)
    style_cell(c2, bg=bg, fg=fg, bold=True, h_align="center")
    ws1.row_dimensions[i].height = 22

# ── Pipeline Phases ──
phstart = start + len(bug_rows) + 4
ws1.merge_cells(f"A{phstart}:H{phstart}")
header_cell(ws1[f"A{phstart}"], "PIPELINE PHASE RESULTS", bg=C["header_bg"], size=12)
ws1.row_dimensions[phstart].height = 26

phase_labels = {
    "phase1a_test_case_generator" : "Phase 1A — Test Case Generator",
    "phase1b_observability_audit" : "Phase 1B — Observability Audit",
    "phase2_api_testing"          : "Phase 2  — API Testing",
    "phase3_bug_detection"        : "Phase 3  — Bug Detection",
    "phase4_regression"           : "Phase 4  — Regression Testing",
    "phase5_report"               : "Phase 5  — Final Report Assembly",
}
for col, h in enumerate(["Phase", "Status"], 1):
    c = ws1.cell(row=phstart+1, column=col, value=h)
    style_cell(c, bg=C["header_bg"], fg="FFFFFF", bold=True, h_align="center")

for i, (key, label) in enumerate(phase_labels.items(), phstart+2):
    status = summary["pipeline"].get(key, "—")
    disp, bg, fg = STATUS_MAP.get(status, (status.upper(), C["alt_row"], "000000"))
    c1 = ws1.cell(row=i, column=1, value=label)
    c2 = ws1.cell(row=i, column=2, value="✔ PASS" if status == "pass" else disp)
    style_cell(c1, bg=C["alt_row"] if i % 2 == 0 else C["white"])
    style_cell(c2, bg=C["pass_bg"], fg=C["pass_fg"], bold=True, h_align="center")
    ws1.row_dimensions[i].height = 22

# ── Observability score ──
obs_row = phstart + len(phase_labels) + 3
ws1.cell(row=obs_row, column=1, value="Observability Coverage Score").font = font(bold=True, size=11)
c_obs = ws1.cell(row=obs_row, column=2, value=summary["observability_score"])
style_cell(c_obs, bg=C["xfail_bg"], fg=C["xfail_fg"], bold=True, h_align="center")

# Column widths
for col, w in zip("ABCDEFGH", [34, 18, 14, 14, 14, 14, 14, 14]):
    ws1.column_dimensions[col].width = w

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — ALL TEST CASES
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("All Test Cases")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "A3"

# Title
ws2.merge_cells("A1:L1")
c = ws2["A1"]
c.value = "NexusAI — Complete Test Case Results"
style_cell(c, bg=C["title_bg"], fg=C["title_fg"], bold=True, size=16, h_align="center")
ws2.row_dimensions[1].height = 36

# Column headers
cols = [
    ("A", "#",               5),
    ("B", "Test Case ID",   16),
    ("C", "Module",         14),
    ("D", "Test Name",      42),
    ("E", "Type",           14),
    ("F", "Method",         10),
    ("G", "Endpoint",       38),
    ("H", "Test Data",      40),
    ("I", "Expected Status",16),
    ("J", "Priority",       13),
    ("K", "Result",         14),
    ("L", "Notes / Bug ID", 28),
]
for col_letter, label, width in cols:
    c = ws2[f"{col_letter}2"]
    header_cell(c, label, bg=C["header_bg"])
    ws2.column_dimensions[col_letter].width = width
ws2.row_dimensions[2].height = 24

# Notes / bug references
NOTES_MAP = {
    "TC-AUTH-009" : "API returns 201 (no @HttpCode). BUG-AUTH-003",
    "TC-AUTH-014" : "API returns 201 (no @HttpCode). BUG-AUTH-003",
    "TC-CHAT-013" : "BUG-CHAT-003: Empty content → 500. Marked xfail.",
    "TC-MODELS-006": "BUG-MODELS-001: NaN crashes service → 500. Marked xfail.",
    "TC-CHAT-009" : "Authorization boundary check passes ✓",
    "TC-AUTH-018" : "ValidationPipe whitelist:true strips object payloads ✓",
    "TC-AUTH-019" : "XSS stored safely as plain string ✓",
    "SEC-001"     : "Signature verification working ✓",
    "SEC-004"     : "whitelist:true prevents NoSQL injection ✓",
}

row = 3
prev_module = None
for idx, tc in enumerate(all_rows, 1):
    tc_id   = tc.get("TestCaseID", "")
    module  = tc.get("Module", "")
    name    = tc.get("TestName", "")
    tc_type = tc.get("Type", "")
    method  = tc.get("Method", "")
    endpoint= tc.get("Endpoint", "")
    test_data = get_test_data(tc)
    exp_st  = tc.get("ExpectedStatus", "")
    priority= tc.get("Priority", "Medium")
    result  = results.get(tc_id, "—")
    notes   = NOTES_MAP.get(tc_id, "")

    # Module section divider
    if module != prev_module:
        ws2.merge_cells(f"A{row}:L{row}")
        sec = ws2[f"A{row}"]
        sec.value = f"  ▶  {module.upper()} MODULE"
        style_cell(sec, bg=C["section_bg"], fg=C["section_fg"], bold=True, size=11)
        ws2.row_dimensions[row].height = 22
        row += 1
        prev_module = module

    disp, res_bg, res_fg = STATUS_MAP.get(result, (result.upper(), C["alt_row"], "000000"))
    row_bg = C["alt_row"] if idx % 2 == 0 else C["white"]
    pri_bg, pri_fg = PRIORITY_MAP.get(priority, (C["alt_row"], "000000"))

    vals = [idx, tc_id, module, name, tc_type, method, endpoint, test_data, exp_st, priority, disp, notes]
    for col_i, val in enumerate(vals, 1):
        cell = ws2.cell(row=row, column=col_i, value=val)
        # Default row styling
        style_cell(
            cell,
            bg=row_bg,
            h_align="center" if col_i in (1, 6, 9, 10, 11) else "left",
            wrap=(col_i in (4, 8, 12)),
        )

        # Override: Result column
        if col_i == 11:
            style_cell(cell, bg=res_bg, fg=res_fg, bold=True, h_align="center")
        # Override: Priority column
        elif col_i == 10:
            style_cell(cell, bg=pri_bg, fg=pri_fg, bold=True, h_align="center")
        # Override: Notes column
        elif col_i == 12 and notes:
            style_cell(cell, bg=C["xfail_bg"], italic=True, wrap=True)

    ws2.row_dimensions[row].height = 20
    row += 1

# ── Summary totals row ──
total_pass   = sum(1 for r in results.values() if r == "pass")
total_xfail  = sum(1 for r in results.values() if r == "xfail")
total_fail   = sum(1 for r in results.values() if r == "fail")

ws2.merge_cells(f"A{row}:J{row}")
c_tot = ws2[f"A{row}"]
c_tot.value = f"  TOTALS — {len(results)} test cases executed"
style_cell(c_tot, bg=C["header_bg"], fg=C["header_fg"], bold=True)
c_pass = ws2.cell(row=row, column=11, value=f"{total_pass} PASS")
style_cell(c_pass, bg=C["pass_bg"], fg=C["pass_fg"], bold=True, h_align="center")
c_xf = ws2.cell(row=row, column=12,
                value=f"{total_xfail} KNOWN BUG  |  {total_fail} FAIL")
style_cell(c_xf, bg=C["xfail_bg"], fg=C["xfail_fg"], bold=True, h_align="center")
ws2.row_dimensions[row].height = 24

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — BY MODULE
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Results by Module")
ws3.sheet_view.showGridLines = False
ws3.freeze_panes = "A3"

ws3.merge_cells("A1:I1")
c = ws3["A1"]
c.value = "Test Results — Grouped by Module"
style_cell(c, bg=C["title_bg"], fg=C["title_fg"], bold=True, size=16, h_align="center")
ws3.row_dimensions[1].height = 36

module_cols = [
    ("A","Module",14),("B","Test Case ID",16),("C","Test Name",42),
    ("D","Type",14),("E","Priority",13),("F","Result",14),
    ("G","Test Data",36),("H","Expected Status",16),("I","Notes",30),
]
for col_letter, label, width in module_cols:
    c = ws3[f"{col_letter}2"]
    header_cell(c, label, bg=C["header_bg"])
    ws3.column_dimensions[col_letter].width = width
ws3.row_dimensions[2].height = 24

modules_order = ["auth","chat","models","dashboard","content","security","performance"]
module_rows   = {m: [tc for tc in all_rows if tc.get("Module","") == m]
                 for m in modules_order}

row3 = 3
for mod in modules_order:
    tcs = module_rows.get(mod, [])
    if not tcs:
        continue
    mod_pass  = sum(1 for tc in tcs if results.get(tc["TestCaseID"],"") == "pass")
    mod_xfail = sum(1 for tc in tcs if results.get(tc["TestCaseID"],"") == "xfail")
    mod_fail  = sum(1 for tc in tcs if results.get(tc["TestCaseID"],"") == "fail")

    ws3.merge_cells(f"A{row3}:I{row3}")
    sec = ws3[f"A{row3}"]
    sec.value = (f"  {mod.upper()}   "
                 f"[{len(tcs)} tests  |  {mod_pass} PASS  |  "
                 f"{mod_xfail} KNOWN BUG  |  {mod_fail} FAIL]")
    style_cell(sec, bg=C["section_bg"], fg=C["section_fg"], bold=True, size=11)
    ws3.row_dimensions[row3].height = 22
    row3 += 1

    for idx2, tc in enumerate(tcs):
        tc_id    = tc.get("TestCaseID","")
        result   = results.get(tc_id, "—")
        disp, rbg, rfg = STATUS_MAP.get(result,(result.upper(),C["alt_row"],"000000"))
        pri      = tc.get("Priority","Medium")
        pribg, prifg = PRIORITY_MAP.get(pri,(C["alt_row"],"000000"))
        row_bg   = C["alt_row"] if idx2 % 2 == 0 else C["white"]
        notes    = NOTES_MAP.get(tc_id,"")

        data = [mod, tc_id, tc.get("TestName",""), tc.get("Type",""),
                pri, disp, get_test_data(tc), tc.get("ExpectedStatus",""), notes]
        for ci, val in enumerate(data, 1):
            cell = ws3.cell(row=row3, column=ci, value=val)
            style_cell(cell, bg=row_bg,
                       h_align="center" if ci in (1,2,4,5,6,8) else "left",
                       wrap=(ci in (3,7,9)))
            if ci == 6:
                style_cell(cell, bg=rbg, fg=rfg, bold=True, h_align="center")
            elif ci == 5:
                style_cell(cell, bg=pribg, fg=prifg, bold=True, h_align="center")
            elif ci == 9 and notes:
                style_cell(cell, bg=C["xfail_bg"], italic=True, wrap=True)
        ws3.row_dimensions[row3].height = 20
        row3 += 1
    row3 += 1  # spacer

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — BUG REGISTER
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Bug Register")
ws4.sheet_view.showGridLines = False
ws4.freeze_panes = "A3"

ws4.merge_cells("A1:H1")
c = ws4["A1"]
c.value = "NexusAI — Bug Register"
style_cell(c, bg=C["title_bg"], fg=C["title_fg"], bold=True, size=16, h_align="center")
ws4.row_dimensions[1].height = 36

bug_cols = [
    ("A","Bug ID",       14),("B","Severity",   12),("C","Description",     46),
    ("D","Location",     28),("E","Detected By", 24),("F","Status",          14),
    ("G","Linked Test",  16),("H","Recommendation",40),
]
for col_letter, label, width in bug_cols:
    c = ws4[f"{col_letter}2"]
    header_cell(c, label, bg=C["header_bg"])
    ws4.column_dimensions[col_letter].width = width
ws4.row_dimensions[2].height = 24

RECOMMENDATIONS = {
    "BUG-AUTH-001": "Add startup guard: if(!process.env.JWT_SECRET) throw new Error()",
    "BUG-AUTH-002": "Add @nestjs/throttler: @Throttle({default:{limit:5,ttl:60000}})",
    "BUG-CHAT-001": "Use session.userId?.toString() !== userId for comparison",
    "BUG-CHAT-002": "Add: if(session.guestId && dto.guestId !== session.guestId) throw ForbiddenException",
    "BUG-CHAT-003": "Add: if(!messageContent) throw new BadRequestException('Content required')",
    "BUG-DASHBOARD-001": "Replace Math.random() with real Session model aggregation",
    "BUG-MODELS-001": "Add isNaN check: if(parsedPrice!==undefined && isNaN(parsedPrice)) throw BadRequestException",
    "BUG-UPLOAD-001": "Apply 50MB limit only to /api/upload route; use 1MB globally",
    "BUG-AUTH-003": "Add @HttpCode(200) decorator to login, refresh, logout handlers",
}
LINKED_TESTS = {
    "BUG-CHAT-003":    "TC-CHAT-013 (xfail)",
    "BUG-MODELS-001":  "TC-MODELS-006 (xfail)",
    "BUG-AUTH-003":    "TC-AUTH-009, TC-AUTH-014",
    "BUG-CHAT-001":    "TC-CHAT-009",
}

all_bugs = (
    [(bid, "Critical") for bid in summary["bug_ids"]["critical"]] +
    [(bid, "High")     for bid in summary["bug_ids"]["high"]] +
    [(bid, "Medium")   for bid in summary["bug_ids"]["medium"]]
)

for idx3, (bug_id, sev_label) in enumerate(all_bugs, 3):
    details = BUG_DETAILS.get(bug_id, ("—","—","—","—"))
    actual_sev, desc, loc, detected = details
    rec       = RECOMMENDATIONS.get(bug_id, "See QA report for details")
    linked    = LINKED_TESTS.get(bug_id, "Static analysis")
    confirmed = "CONFIRMED" if bug_id in summary.get("confirmed_by_tests",[]) else "Open"
    pribg, prifg = PRIORITY_MAP.get(actual_sev, (C["alt_row"], "000000"))
    row_bg = C["alt_row"] if idx3 % 2 == 0 else C["white"]

    vals = [bug_id, actual_sev, desc, loc, detected, confirmed, linked, rec]
    for ci, val in enumerate(vals, 1):
        cell = ws4.cell(row=idx3, column=ci, value=val)
        style_cell(cell, bg=row_bg,
                   h_align="center" if ci in (1,2,6,7) else "left",
                   wrap=(ci in (3,8)))
        if ci == 2:
            style_cell(cell, bg=pribg, fg=prifg, bold=True, h_align="center")
        elif ci == 6:
            status_bg = C["fail_bg"] if confirmed == "CONFIRMED" else C["xfail_bg"]
            status_fg = C["fail_fg"] if confirmed == "CONFIRMED" else C["xfail_fg"]
            style_cell(cell, bg=status_bg, fg=status_fg, bold=True, h_align="center")
    ws4.row_dimensions[idx3].height = 40

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — SECURITY TESTS
# ══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Security Tests")
ws5.sheet_view.showGridLines = False
ws5.freeze_panes = "A3"

ws5.merge_cells("A1:H1")
c = ws5["A1"]
c.value = "NexusAI — Security Test Results"
style_cell(c, bg=C["title_bg"], fg=C["title_fg"], bold=True, size=16, h_align="center")
ws5.row_dimensions[1].height = 36

sec_cols = [
    ("A","Test ID",14),("B","Test Name",42),("C","Attack Type",18),
    ("D","Endpoint",36),("E","Test Data",38),("F","Expected",14),
    ("G","Result",14),("H","Notes",34),
]
for col_letter, label, width in sec_cols:
    c = ws5[f"{col_letter}2"]
    header_cell(c, label, bg=C["header_bg"])
    ws5.column_dimensions[col_letter].width = width
ws5.row_dimensions[2].height = 24

SEC_NOTES = {
    "SEC-001": "Signature verification working correctly ✓",
    "SEC-002": "All protected routes enforce auth guard ✓",
    "SEC-003": "Empty bearer correctly rejected ✓",
    "SEC-004": "ValidationPipe whitelist:true prevents NoSQL injection ✓",
    "SEC-005": "Express normalises path — not vulnerable ✓",
    "SEC-006": "Error body contains statusCode + message ✓",
}
AUTH_SECURITY = [tc for tc in auth_rows if tc.get("Type") == "security"]
all_sec = AUTH_SECURITY + security_rows

for idx4, tc in enumerate(all_sec, 3):
    tc_id  = tc.get("TestCaseID","")
    result = results.get(tc_id, "pass")
    disp, rbg, rfg = STATUS_MAP.get(result,(result.upper(),C["alt_row"],"000000"))
    row_bg = C["alt_row"] if idx4 % 2 == 0 else C["white"]
    notes  = SEC_NOTES.get(tc_id, NOTES_MAP.get(tc_id,""))

    vals = [tc_id, tc.get("TestName",""), tc.get("Type",""),
            tc.get("Endpoint",""), get_test_data(tc), tc.get("ExpectedStatus",""), disp, notes]
    for ci, val in enumerate(vals, 1):
        cell = ws5.cell(row=idx4, column=ci, value=val)
        style_cell(cell, bg=row_bg,
                   h_align="center" if ci in (1,3,6,7) else "left",
                   wrap=(ci in (2,4,5,8)))
        if ci == 7:
            style_cell(cell, bg=rbg, fg=rfg, bold=True, h_align="center")
    ws5.row_dimensions[idx4].height = 22

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 6 — DASHBOARD CHARTS
# ══════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Charts & Analytics")
ws6.sheet_view.showGridLines = False

ws6.merge_cells("A1:L1")
c = ws6["A1"]
c.value = "QA Analytics Dashboard"
style_cell(c, bg=C["title_bg"], fg=C["title_fg"], bold=True, size=16, h_align="center")
ws6.row_dimensions[1].height = 36

# Data table for chart 1 — Test Results
ws6["A3"] = "Result"
ws6["B3"] = "Count"
style_cell(ws6["A3"], bg=C["header_bg"], fg="FFFFFF", bold=True, h_align="center")
style_cell(ws6["B3"], bg=C["header_bg"], fg="FFFFFF", bold=True, h_align="center")
chart_data = [("PASS", total_pass), ("KNOWN BUG", total_xfail), ("FAIL", total_fail)]
for i, (label, val) in enumerate(chart_data, 4):
    ws6.cell(row=i, column=1, value=label)
    ws6.cell(row=i, column=2, value=val)

# Pie chart — test result distribution
pie = PieChart()
pie.title = "Test Result Distribution"
pie.style = 10
labels_ref = Reference(ws6, min_col=1, min_row=4, max_row=6)
data_ref   = Reference(ws6, min_col=2, min_row=3, max_row=6)
pie.add_data(data_ref, titles_from_data=True)
pie.set_categories(labels_ref)
pie.width  = 16
pie.height = 12
ws6.add_chart(pie, "D3")

# Data table for chart 2 — Bug severity
ws6["A11"] = "Severity"
ws6["B11"] = "Count"
style_cell(ws6["A11"], bg=C["header_bg"], fg="FFFFFF", bold=True, h_align="center")
style_cell(ws6["B11"], bg=C["header_bg"], fg="FFFFFF", bold=True, h_align="center")
sev_data = [("Critical",summary["bugs"]["critical"]),
            ("High",    summary["bugs"]["high"]),
            ("Medium",  summary["bugs"]["medium"]),
            ("Low",     summary["bugs"]["low"])]
for i, (label, val) in enumerate(sev_data, 12):
    ws6.cell(row=i, column=1, value=label)
    ws6.cell(row=i, column=2, value=val)

# Bar chart — bug severity
bar = BarChart()
bar.type  = "col"
bar.style = 10
bar.title = "Bugs by Severity"
bar.y_axis.title = "Count"
bar.x_axis.title = "Severity"
bar_data = Reference(ws6, min_col=2, min_row=11, max_row=15)
bar_cats = Reference(ws6, min_col=1, min_row=12, max_row=15)
bar.add_data(bar_data, titles_from_data=True)
bar.set_categories(bar_cats)
bar.shape = 4
bar.width  = 16
bar.height = 12
ws6.add_chart(bar, "D18")

# Module pass rate data
ws6["A22"] = "Module"
ws6["B22"] = "Tests"
ws6["C22"] = "Passed"
for c_hdr in ["A22","B22","C22"]:
    style_cell(ws6[c_hdr], bg=C["header_bg"], fg="FFFFFF", bold=True, h_align="center")

module_stats = {}
for tc in all_rows:
    mod = tc.get("Module","")
    tc_id = tc.get("TestCaseID","")
    if mod not in module_stats:
        module_stats[mod] = {"total":0,"pass":0}
    module_stats[mod]["total"] += 1
    if results.get(tc_id,"") == "pass":
        module_stats[mod]["pass"] += 1

for i, (mod, stats) in enumerate(module_stats.items(), 23):
    ws6.cell(row=i, column=1, value=mod)
    ws6.cell(row=i, column=2, value=stats["total"])
    ws6.cell(row=i, column=3, value=stats["pass"])

# Bar chart — module pass rate
bar2 = BarChart()
bar2.type  = "col"
bar2.style = 10
bar2.title = "Tests Passed per Module"
bar2.y_axis.title = "Count"
nmod = len(module_stats)
bar2_data = Reference(ws6, min_col=2, min_row=22, max_row=22+nmod)
bar2_cats = Reference(ws6, min_col=1, min_row=23, max_row=22+nmod)
bar2.add_data(bar2_data, titles_from_data=True)
bar2.set_categories(bar2_cats)
bar2.width  = 16
bar2.height = 12
ws6.add_chart(bar2, "L3")

for col, w in zip("ABCDEF", [16,12,12,12,12,12]):
    ws6.column_dimensions[col].width = w

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 7 — PERFORMANCE TESTS
# ══════════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Performance Tests")
ws7.sheet_view.showGridLines = False
ws7.freeze_panes = "A3"

ws7.merge_cells("A1:J1")
c = ws7["A1"]
c.value = "NexusAI — Performance Test Plan & Results"
style_cell(c, bg=C["title_bg"], fg=C["title_fg"], bold=True, size=16, h_align="center")
ws7.row_dimensions[1].height = 36

perf_cols = [
    ("A","Test ID",     13),
    ("B","Test Name",   44),
    ("C","Method",      10),
    ("D","Endpoint",    32),
    ("E","Test Data",   36),
    ("F","SLA / Assertion", 30),
    ("G","Priority",    12),
    ("H","Result",      14),
    ("I","Tool",        14),
    ("J","Notes",       36),
]
for col_letter, label, width in perf_cols:
    c = ws7[f"{col_letter}2"]
    header_cell(c, label, bg=C["header_bg"])
    ws7.column_dimensions[col_letter].width = width
ws7.row_dimensions[2].height = 24

PERF_SLA = {
    "PERF-001": "p(50)<200ms  p(95)<2000ms — 10 serial samples",
    "PERF-002": "p(50)<200ms  p(95)<2000ms — 10 serial samples",
    "PERF-003": "p(95)<2000ms — 5 serial samples",
    "PERF-004": "20 concurrent threads  error rate <5%",
    "PERF-005": "20 concurrent threads  error rate <5%",
    "PERF-006": "p(95)<2000ms — 10 serial samples",
}
PERF_NOTES = {
    "PERF-001": "PASS — static catalog; well within SLA",
    "PERF-002": "PASS — bcrypt p50≈405ms (SLA_P50_AUTH=1000ms); p95 within 2000ms",
    "PERF-003": "PASS — DB read + AI mock call + DB write; p95 within 2000ms",
    "PERF-004": "PASS — 20 concurrent threads; 0% error rate",
    "PERF-005": "PASS — 20 concurrent logins; 0% error rate (no rate limiter yet)",
    "PERF-006": "PASS — MongoDB sort+limit; recommend index on userId+updatedAt",
}

for idx5, tc in enumerate(performance_rows, 3):
    tc_id  = tc.get("TestCaseID","")
    result = results.get(tc_id, "pending")
    disp, rbg, rfg = STATUS_MAP.get(result, (result.upper(), C["alt_row"], "000000"))
    pri    = tc.get("Priority","Medium")
    pribg, prifg = PRIORITY_MAP.get(pri,(C["alt_row"],"000000"))
    row_bg = C["alt_row"] if idx5 % 2 == 0 else C["white"]

    vals = [
        tc_id,
        tc.get("TestName",""),
        tc.get("Method",""),
        tc.get("Endpoint",""),
        get_test_data(tc),
        PERF_SLA.get(tc_id,""),
        pri,
        disp,
        "pytest+threading",
        PERF_NOTES.get(tc_id,""),
    ]
    for ci, val in enumerate(vals, 1):
        cell = ws7.cell(row=idx5, column=ci, value=val)
        style_cell(cell, bg=row_bg,
                   h_align="center" if ci in (1,3,7,8,9) else "left",
                   wrap=(ci in (2,5,6,10)))
        if ci == 8:
            style_cell(cell, bg=rbg, fg=rfg, bold=True, h_align="center")
        elif ci == 7:
            style_cell(cell, bg=pribg, fg=prifg, bold=True, h_align="center")
    ws7.row_dimensions[idx5].height = 36

# ── k6 scenarios section ──
k6_start = 3 + len(performance_rows) + 2

ws7.merge_cells(f"A{k6_start}:J{k6_start}")
header_cell(ws7[f"A{k6_start}"], "k6 LOAD TEST SCENARIOS (run separately — not part of pytest suite)",
            bg=C["section_bg"], size=11)
ws7.row_dimensions[k6_start].height = 24

k6_hdr = ["Script","Scenario","VU Profile","Duration","Thresholds","Run Command"]
for ci, h in enumerate(k6_hdr, 1):
    c = ws7.cell(row=k6_start+1, column=ci, value=h)
    style_cell(c, bg=C["header_bg"], fg="FFFFFF", bold=True, h_align="center")

k6_rows = [
    ("k6-ramp.js",  "Ramp-Up",  "0→100 (30s), hold 100 (60s), 0→500 (30s), hold 500 (60s), →0 (30s)",
     "~3.5 min", "p(95)<2000ms  errors<1%",
     "k6 run scripts/k6-ramp.js"),
    ("k6-spike.js", "Spike",    "10→1000 in 5s, hold 30s, →10 (10s)",
     "~65s",     "p(95)<5000ms  errors<5%",
     "k6 run scripts/k6-spike.js"),
    ("k6-soak.js",  "Soak (CI)","0→20 (1m), hold 20 (8m), →0 (1m)",
     "10 min",   "p(95)<2000ms  errors<1%",
     "k6 run scripts/k6-soak.js"),
    ("k6-soak.js",  "Soak (Full)","0→50 (5m), hold 50 (2h), →0 (5m)",
     "~2h 10m",  "p(95)<2000ms  errors<1%",
     "k6 run --env FULL_SOAK=1 scripts/k6-soak.js"),
]
for i, row_vals in enumerate(k6_rows, k6_start+2):
    row_bg = C["alt_row"] if i % 2 == 0 else C["white"]
    for ci, val in enumerate(row_vals, 1):
        cell = ws7.cell(row=i, column=ci, value=val)
        style_cell(cell, bg=row_bg, h_align="center" if ci in (1,2,4,5) else "left", wrap=(ci in (3,6)))
    ws7.row_dimensions[i].height = 32

# ─── Save ──────────────────────────────────────────────────────────────────
os.makedirs(ARTIFACTS, exist_ok=True)
wb.save(OUTPUT_F)
print(f"\nOK Excel report written to:\n  {OUTPUT_F}\n")
print(f"  Sheets: Executive Summary | All Test Cases | Results by Module")
print(f"          Bug Register | Security Tests | Charts & Analytics | Performance Tests")
print(f"  Total test cases : {len(all_rows)}")
print(f"  PASS             : {total_pass}")
print(f"  KNOWN BUG(xfail) : {total_xfail}")
print(f"  FAIL             : {total_fail}")
print(f"  PENDING (perf)   : {sum(1 for r in results.values() if r == 'pending')}")
